#Password Keeper Commander

__author__ = "Shreyank Ramachandra"
__email__ = "shreyankkr@gmail.com"
__version__ = 0.1

import getpass
import os
import sys
import argparse
import re
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

from keepercommander.api import login, search_records, get_record_shares, sync_down, get_record, update_record
from keepercommander.params import KeeperParams
from keepercommander.generator import generate


def get_user_pass():
    '''Function to get Keeper Security's username and password to login
    Returns:
        params (keepercommander.params.KeeperParams): Used as Global parameter and session container
    '''
    keeper_email = input("Enter your Keeper Security vault Email : ")
    keeper_pass = getpass.getpass(prompt=f"Enter Keeper Security vault Password for {keeper_email} : ", stream=None)
    
    params = KeeperParams()
    params.user = keeper_email
    params.password = keeper_pass
    try:
        login(params)
    except Exception as e:
        print(e)
        print("Retry Again with correct username ans password...")
        sys.exit()
    if params.session_token:
        print("Successfully Logged in to Keeper Security")
        sync_down(params)
        return params


def get_all_keeper_records(params):
    '''This runs keepercommander's list command and get username and passwords
    Args:
        params (keepercommander.params.KeeperParams): Used as Global parameter and session container
    Returns:
        results (list) : list of login records
    '''
    params.commands= ['list']
    results = search_records(params, '')
    if results:
        return results
    else:
        print(f"Unable to fetch any records from Keeper Security for {params.user}")
        sys.exit()
    

def rotate_web_pass(record, params, driver):
    '''Function to Login and change the password in the portal
    Args:
        record (keepercommander.record.Record): Record object
        params (keepercommander.params.KeeperParams): Used as Global parameter and session container
        driver (Webdriver) : selenium webdriver object
    Returns:
        stat (str) : Status of changing password
    '''
    print (rf'Rotating Password for "{record.title}" in {record.login_url[:50]}...')
    driver.get(record.login_url)
    
    # TODO : Login to your site here...

    # TODO : Navigate to Reset password Page here... 

    # Generate Random Passowrd as below... 
    new_gen_pass = generate(16)
    
    # TODO : Change in the reset password page in website 
    # If success in website update record in Keeper Security as below...
    record.password = new_gen_pass
    if update_record(params, record):
        stat = f"Successfully Changed password in {record.login_url[:60]}..."
        print(stat)
    else:
        # Incase there is network issue!
        stat=f"Unable to save new password in Keeper Security. {record.title} new password is {new_gen_pass}. Please change manually in Keeper Security Vault"
        print(stat)
    return stat


if __name__ == "__main__":
    """Main function for the program"""
    parser = argparse.ArgumentParser(description="Password Rotation Bot using keepercommander")
    parser.add_argument('-x', '--xls_uid', dest='xls_uid', default='rotate_uids.xlsx', help="Path for excel file with 2nd column having Keeper Security Record UID to rotate the passwords")

    args = parser.parse_args()
    uid_excel = os.path.abspath(args.xls_uid)
    if os.path.exists(uid_excel):
        print("Password Rotation Bot for Keeper Security Vault")
        print("---------------------------------------------------------------")
        params = get_user_pass()
        print(f"Reading data from {args.xls_uid}...")
        data_in = load_workbook(args.xls_uid)
        # Get first sheet in workbook
        input_sheet = data_in[data_in.get_sheet_names()[0]]
        logins_list = get_all_keeper_records(params)
        # Open Chrome to change password in the website of interest
        driver = webdriver.Chrome() # chromedriver.exe is in system PATH or in this script's path
        # For Each UID in Excel input rotate passwords
        for rows in range(3 , input_sheet.max_row+1):
            uid = input_sheet.cell(rows, 2).value # UIDs are in 2nd columns
            uid_rec = [each_rec for each_rec in logins_list if re.search(uid, each_rec.record_uid, re.I)] # Get the record details from Keeper Security
            if uid_rec:
                uid_rec = uid_rec[0]
                print("\n" + f'Rotating for {uid_rec.record_uid} with login "{uid_rec.login}":')
                change_status = rotate_web_pass(uid_rec, params, driver) #Change in Website and then update in Keeper Security Vault
                #Update results in the same excel
                input_sheet.cell(rows, 5, change_status)
                data_in.save(args.xls_uid)
                
        input("\nRotating Passwords completed. Press enter to exit")
        driver.quit()
    else:
        sys.exit(f"Unable to find the Input file : {uid_excel}. Exiting application...")